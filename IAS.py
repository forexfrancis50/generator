import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from io import BytesIO
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Function to calculate DCF
def calculate_dcf(user_inputs):
    """
    Calculate DCF based on provided inputs.
    """
    tax_rate = user_inputs["tax_rate"]
    long_term_growth_rate = user_inputs["long_term_growth_rate"]
    wacc = user_inputs["wacc"]
    share_price = user_inputs["share_price"]
    shares_outstanding = user_inputs["shares_outstanding"]
    ebit = user_inputs["ebit"]
    da = user_inputs["da"]
    capex = user_inputs["capex"]
    changes_in_nwc = user_inputs["changes_in_nwc"]
    cash = user_inputs["cash"]
    debt = user_inputs["debt"]

    # Calculate Free Cash Flow (FCF) for each period
    fcf = []
    for i in range(5):
        tax_on_ebit = ebit[i] * tax_rate * -1
        fcf.append(ebit[i] + tax_on_ebit + da[i] + capex[i] + changes_in_nwc[i])

    # Calculate Terminal Value
    terminal_value = fcf[-1] * (1 + long_term_growth_rate) / (wacc - long_term_growth_rate)

    # Discount Cash Flows
    present_values = []
    for i in range(5):
        discount_factor = 1 / (1 + wacc) ** (i + 1)
        present_values.append(fcf[i] * discount_factor)

    # Discount Terminal Value
    discount_factor_terminal = 1 / (1 + wacc) ** 5
    present_terminal_value = terminal_value * discount_factor_terminal

    # Calculate Enterprise Value
    enterprise_value = sum(present_values) + present_terminal_value

    # Calculate Equity Value
    equity_value = enterprise_value + cash - debt

    # Calculate Intrinsic Value per Share
    intrinsic_value_per_share = equity_value / shares_outstanding

    # Calculate Market Capitalization
    market_cap = share_price * shares_outstanding

    # Calculate Intrinsic Value Premium
    intrinsic_value_premium = equity_value - market_cap
    intrinsic_value_premium_percentage = intrinsic_value_premium / market_cap

    return {
        "fcf": fcf,
        "terminal_value": terminal_value,
        "present_values": present_values,
        "present_terminal_value": present_terminal_value,
        "enterprise_value": enterprise_value,
        "equity_value": equity_value,
        "intrinsic_value_per_share": intrinsic_value_per_share,
        "market_cap": market_cap,
        "intrinsic_value_premium": intrinsic_value_premium,
        "intrinsic_value_premium_percentage": intrinsic_value_premium_percentage,
    }

# Function to apply professional formatting to the spreadsheet
def apply_formatting(ws):
    """
    Apply professional formatting to the worksheet.
    """
    # Set column widths
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Set header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_border = Border(bottom=Side(border_style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=4, max_col=12):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center")

    # Set currency formatting
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=12):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0.00'

    # Set percentage formatting for specific cells
    percentage_cells = ['B5', 'B6', 'B7']  # Tax Rate, Discount Rate, Perpetual Growth Rate
    for cell_ref in percentage_cells:
        cell = ws[cell_ref]
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0.00%'

    # Add borders to data
    thin_border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=12):
        for cell in row:
            cell.border = thin_border

# Function to add charts to the spreadsheet
def add_charts(ws, dcf_results):
    """
    Add charts to the worksheet.
    """
    # Line chart for cash flows
    chart = LineChart()
    chart.title = "Unlevered Free Cash Flow (UFCF)"
    chart.style = 13
    chart.y_axis.title = "Cash Flow ($)"
    chart.x_axis.title = "Year"
    data = Reference(ws, min_col=6, min_row=13, max_col=10, max_row=13)
    categories = Reference(ws, min_col=5, min_row=12, max_col=10, max_row=12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "A30")

# Function to generate Excel spreadsheet
def generate_spreadsheet(user_inputs, dcf_results, company_name):
    """
    Generate an Excel spreadsheet with DCF results.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Model"

    # Write headers
    ws.append([f"DCF Model for {company_name}"])
    ws.append(["All figures in millions unless otherwise stated"])
    ws.append([])
    ws.append(["", "Hist.", "Proj.", "Proj.", "Proj.", "Proj.", "Proj."])
    ws.append(["", "Period 0", "Period 1", "Period 2", "Period 3", "Period 4", "Period 5"])
    ws.append([])
    ws.append(["Discounted Cash Flow"])
    ws.append(["", "Tax rate", user_inputs["tax_rate"]])
    ws.append(["", "Long term growth rate", user_inputs["long_term_growth_rate"]])
    ws.append(["", "WACC", user_inputs["wacc"]])
    ws.append(["", "Share price", user_inputs["share_price"]])
    ws.append(["", "Shares outstanding", user_inputs["shares_outstanding"]])
    ws.append([])
    ws.append(["", "Year count", "", 1, 2, 3, 4, 5])
    ws.append(["", "EBIT", "", *user_inputs["ebit"]])
    ws.append(["", "Tax on EBIT", "", *[ebit * user_inputs["tax_rate"] * -1 for ebit in user_inputs["ebit"]]])
    ws.append(["", "+ Depreciation and amortization", "", *user_inputs["da"]])
    ws.append(["", "- Capital expenditure", "", *user_inputs["capex"]])
    ws.append(["", "Change in operating working capital", "", *user_inputs["changes_in_nwc"]])
    ws.append(["", "Free cash flow", "", *dcf_results["fcf"]])
    ws.append([])
    ws.append(["", "Terminal value", "", "", "", "", "", dcf_results["terminal_value"]])
    ws.append([])
    ws.append(["", "Discount factor", "", *[1 / (1 + user_inputs["wacc"]) ** (i + 1) for i in range(5)]])
    ws.append(["", "Present value of free cash flows", "", *dcf_results["present_values"]])
    ws.append([])
    ws.append(["", "Sum of present value of free cash flows", sum(dcf_results["present_values"])])
    ws.append(["", "Present value of terminal value", dcf_results["present_terminal_value"]])
    ws.append(["", "Enterprise value", dcf_results["enterprise_value"]])
    ws.append([])
    ws.append(["", "+ Cash", user_inputs["cash"]])
    ws.append(["", "- Debt", user_inputs["debt"]])
    ws.append(["", "Implied equity value (intrinsic value)", dcf_results["equity_value"]])
    ws.append([])
    ws.append(["", "Market capitalization", dcf_results["market_cap"]])
    ws.append(["", "Intrinsic value premium to market capitalization", dcf_results["intrinsic_value_premium"]])
    ws.append(["", "Intrinsic value premium percentage", dcf_results["intrinsic_value_premium_percentage"]])

    # Apply professional formatting
    apply_formatting(ws)

    # Save to BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def format_currency(value):
    """
    Format a number as currency with commas and 2 decimal places.
    """
    return f"${value:,.2f}"

def format_percentage(value):
    """
    Format a number as percentage with 2 decimal places.
    """
    return f"{value:.2%}"

def display_colored_value(label, value, color, is_percentage=False):
    """
    Display a value with a colored label using Streamlit markdown.
    
    Args:
        label (str): The label to display
        value (float): The numerical value to display
        color (str): The color to use ('red' or 'green')
        is_percentage (bool): Whether to format as percentage
    """
    formatted_value = format_percentage(value) if is_percentage else format_currency(value)
    st.markdown(f"**{label}:** <span style='color: {color}'>{formatted_value}</span>", unsafe_allow_html=True)

def generate_pdf(user_inputs, dcf_results, company_name):
    """
    Generate a PDF report with DCF results.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30
    )
    elements.append(Paragraph(f"DCF Analysis Report for {company_name}", title_style))
    elements.append(Spacer(1, 12))

    # Key Assumptions Table
    assumptions_data = [
        ["Key Assumptions", "Value"],
        ["Tax Rate", f"{user_inputs['tax_rate']:.1%}"],
        ["WACC", f"{user_inputs['wacc']:.1%}"],
        ["Long Term Growth Rate", f"{user_inputs['long_term_growth_rate']:.1%}"],
        ["Share Price", f"${user_inputs['share_price']:.2f}"],
        ["Shares Outstanding", f"{user_inputs['shares_outstanding']:,.0f}"],
    ]
    
    assumptions_table = Table(assumptions_data, colWidths=[3*inch, 2*inch])
    assumptions_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
    ]))
    elements.append(assumptions_table)
    elements.append(Spacer(1, 20))

    # Results Table
    results_data = [
        ["DCF Results", "Value"],
        ["Enterprise Value", f"${dcf_results['enterprise_value']:,.2f}"],
        ["Equity Value", f"${dcf_results['equity_value']:,.2f}"],
        ["Intrinsic Value per Share", f"${dcf_results['intrinsic_value_per_share']:,.2f}"],
        ["Market Cap", f"${dcf_results['market_cap']:,.2f}"],
        ["Intrinsic Value Premium", f"${dcf_results['intrinsic_value_premium']:,.2f}"],
        ["Premium Percentage", f"{dcf_results['intrinsic_value_premium_percentage']:.1%}"],
    ]
    
    results_table = Table(results_data, colWidths=[3*inch, 2*inch])
    results_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
    ]))
    elements.append(results_table)
    
    # Footer
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey
    )
    elements.append(Paragraph("Disclaimer: This report is for educational purposes only.", footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

# Streamlit App
def main():
    st.title("DCF Model Generator")
    st.write("Input your financial data to generate a DCF spreadsheet.")

    # Company Name Input
    company_name = st.text_input("Company Name", value="Example Company", help="Enter the name of the company you are modeling.")
    st.markdown("---")

    # Key Rates and Assumptions
    with st.expander("Key Rates & Assumptions", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            tax_rate = st.number_input("Tax Rate", value=0.21, help="The effective tax rate applied to EBIT.")
            long_term_growth_rate = st.number_input("Long Term Growth Rate", value=0.03, help="The growth rate used to calculate terminal value.")
            wacc = st.number_input("WACC", value=0.08, help="The weighted average cost of capital.")
        with col2:
            share_price = st.number_input("Share Price", value=3.0, help="The current market price per share.")
            shares_outstanding = st.number_input("Shares Outstanding", value=1000.0, help="The total number of shares outstanding.")

    # Company Financials
    with st.expander("Company Financials", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            cash = st.number_input("Cash", value=124.2, help="The total cash and cash equivalents.")
            debt = st.number_input("Debt", value=500.0, help="The total debt of the company.")

    # Yearly Projections
    with st.expander("5-Year Projections", expanded=True):
        # EBIT Projections
        st.subheader("EBIT Projections")
        ebit_cols = st.columns(5)
        ebit = []
        for i, col in enumerate(ebit_cols):
            with col:
                ebit.append(st.number_input(f"Year {i+1}", value=270.0 if i == 0 else 0.0, key=f"ebit_{i}"))

        # D&A Projections
        st.subheader("Depreciation & Amortization")
        da_cols = st.columns(5)
        da = []
        for i, col in enumerate(da_cols):
            with col:
                da.append(st.number_input(f"Year {i+1}", value=20.0 if i == 0 else 0.0, key=f"da_{i}"))

        # CapEx Projections
        st.subheader("Capital Expenditure")
        capex_cols = st.columns(5)
        capex = []
        for i, col in enumerate(capex_cols):
            with col:
                capex.append(st.number_input(f"Year {i+1}", value=-30.0 if i == 0 else 0.0, key=f"capex_{i}"))

        # NWC Changes Projections
        st.subheader("Changes in Net Working Capital")
        nwc_cols = st.columns(5)
        changes_in_nwc = []
        for i, col in enumerate(nwc_cols):
            with col:
                changes_in_nwc.append(st.number_input(f"Year {i+1}", value=-2.8 if i == 0 else 0.0, key=f"nwc_{i}"))

    # Store user inputs in a dictionary
    user_inputs = {
        "tax_rate": tax_rate,
        "long_term_growth_rate": long_term_growth_rate,
        "wacc": wacc,
        "share_price": share_price,
        "shares_outstanding": shares_outstanding,
        "cash": cash,
        "debt": debt,
        "ebit": ebit,
        "da": da,
        "capex": capex,
        "changes_in_nwc": changes_in_nwc,
    }

    # Calculate DCF
    if st.button("Calculate DCF"):
        try:
            dcf_results = calculate_dcf(user_inputs)

            # Display results in a more organized way
            with st.expander("DCF Results", expanded=True):
                st.subheader("DCF Results")
                with st.container():
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        display_colored_value("Enterprise Value", dcf_results["enterprise_value"], "green")
                        display_colored_value("Equity Value", dcf_results["equity_value"], "green")
                    with col2:
                        display_colored_value("Intrinsic Value per Share", dcf_results["intrinsic_value_per_share"], "green")
                        display_colored_value("Market Cap", dcf_results["market_cap"], "blue")
                    with col3:
                        display_colored_value("Intrinsic Value Premium", dcf_results["intrinsic_value_premium"], 
                                           "green" if dcf_results["intrinsic_value_premium"] >= 0 else "red")
                        display_colored_value("Intrinsic Value Premium %", 
                                           dcf_results["intrinsic_value_premium_percentage"], 
                                           "green" if dcf_results["intrinsic_value_premium"] >= 0 else "red",
                                           is_percentage=True)

                # Download buttons in columns
                st.subheader("Download Reports")
                col1, col2 = st.columns(2)
                
                with col1:
                    # Excel download
                    excel_file = generate_spreadsheet(user_inputs, dcf_results, company_name)
                    st.download_button(
                        label="Download Excel File",
                        data=excel_file,
                        file_name=f"dcf_analysis_{company_name.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                
                with col2:
                    # PDF download
                    pdf_file = generate_pdf(user_inputs, dcf_results, company_name)
                    st.download_button(
                        label="Download PDF Report",
                        data=pdf_file,
                        file_name=f"dcf_analysis_{company_name.replace(' ', '_')}.pdf",
                        mime="application/pdf",
                    )

        except Exception as e:
            st.error(f"Error calculating DCF: {e}")

    # Footer
    st.markdown("---")
    st.markdown("**App Version:** 1.0 | **Disclaimer:** This app is for educational purposes only.")

# Run the app
if __name__ == "__main__":
    main()